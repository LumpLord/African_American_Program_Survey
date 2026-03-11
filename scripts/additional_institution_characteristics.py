"""\
Draft scraper for additional institution characteristics from NCES IPEDS Institution Profile pages.

Target fields (initial draft; may need iterative tuning):
1) Tuition and required fees for full-time students (Academic year 2024-25)
   - Undergraduate
   - Graduate
2) Enrollment by gender, student level, and full- and part-time status
   - Total students
   - Total men
   - Total women
3) Percent of all students enrolled, by race/ethnicity
   - American Indian or Alaska Native
   - Asian
   - Black or African American
   - Hispanic
   - Native Hawaiian or Other Pacific Islander
   - White
   - Two or more races
   - Race/ethnicity unknown
   - U.S. Nonresident

This is a TEST-ONLY script:
- Reads a definable head from the input file.
- Prints what it finds.
- DOES NOT write outputs yet.

Intended to be run inside an ipynb notebook.
"""

from __future__ import annotations

from pathlib import Path
import time
import re
from typing import Dict, Optional, Tuple

import pandas as pd

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait

from bs4 import BeautifulSoup

import json
import tempfile
import shutil
import os
import openpyxl


# ----------------------------
# Inputs / knobs
# ----------------------------
INPUT_PATH = Path(
    "ace_unitid_merge__ace_x_2013comp__webscrape__v15simple__bucketed_programs__2013_current_matches.csv"
)

OUTPUT_SUFFIX = "__nces_profile_characteristics"

UNITID_COL = "unitid"
NAME_COL = "name"

NCES_PROFILE_URL = "https://nces.ed.gov/ipeds/institution-profile/{}"
NCES_PROFILE_URL_ANCHOR = "https://nces.ed.gov/ipeds/institution-profile/{}/#"

WAIT_TIMEOUT_SEC = 25
MIN_SLEEP_AFTER_GET_SEC = 0.8
SLEEP_BETWEEN_REQUESTS_SEC = 1.0

TEST_N = 0  # <--- during testing; set to 0 for a full run over all unique unitids


# Debug knob for XLSX discovery
DEBUG_XLSX_DISCOVERY = True
XLSX_CLICK_AND_CAPTURE = True
XLSX_CAPTURE_WINDOW_SEC = 6.0

# XLSX-first workflow knobs
USE_XLSX_EXTRACTION = True
DELETE_DOWNLOADED_XLSX = False  # keep downloaded XLSX files for debugging
DOWNLOAD_WAIT_SEC = 35


# ----------------------------
# Selenium
# ----------------------------


def _ensure_download_dir() -> Path:
    d = Path(tempfile.gettempdir()) / "nces_ipeds_profile_downloads"
    d.mkdir(parents=True, exist_ok=True)
    return d


def make_driver() -> webdriver.Chrome:
    options = webdriver.ChromeOptions()
    options.add_argument("--headless=new")
    options.add_argument("--disable-gpu")
    options.add_argument("--no-sandbox")
    options.add_argument("--window-size=1400,900")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--blink-settings=imagesEnabled=false")

    # Enable Chrome performance logs (kept for optional debugging).
    options.set_capability("goog:loggingPrefs", {"performance": "ALL"})

    driver = webdriver.Chrome(options=options)

    # Enable Network via CDP (optional; helps with debugging)
    try:
        driver.execute_cdp_cmd("Network.enable", {})
    except Exception:
        pass

    # Allow downloads in headless Chrome to a known temp directory.
    dl_dir = _ensure_download_dir()
    try:
        driver.execute_cdp_cmd(
            "Page.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": str(dl_dir)},
        )
    except Exception:
        # Some Chrome/Selenium combos use Browser.setDownloadBehavior
        try:
            driver.execute_cdp_cmd(
                "Browser.setDownloadBehavior",
                {"behavior": "allow", "downloadPath": str(dl_dir)},
            )
        except Exception:
            pass

    # stash for later
    driver._download_dir = dl_dir  # type: ignore[attr-defined]

    return driver


def _drain_performance_logs(driver: webdriver.Chrome):
    """Return and clear performance logs (best-effort)."""
    try:
        return driver.get_log("performance")
    except Exception:
        return []


def _extract_xlsx_like_urls_from_perf(perf_logs) -> list[str]:
    """Parse Chrome performance logs and return candidate URLs that look like XLSX exports."""
    cands: list[str] = []

    for entry in perf_logs or []:
        try:
            msg = json.loads(entry.get("message", "{}")).get("message", {})
        except Exception:
            continue

        method = msg.get("method", "")
        params = msg.get("params", {})

        url = ""
        mime = ""

        if method == "Network.requestWillBeSent":
            req = params.get("request", {})
            url = (req.get("url") or "").strip()

        elif method == "Network.responseReceived":
            resp = params.get("response", {})
            url = (resp.get("url") or "").strip()
            mime = (resp.get("mimeType") or "").strip().lower()

        if not url:
            continue

        ul = url.lower()

        # Direct .xlsx
        if ".xlsx" in ul:
            cands.append(url)
            continue

        # Common export endpoints or query flags even without .xlsx
        if any(tok in ul for tok in ("export", "download", "excel", "xlsx", "format=xls", "format=xlsx")):
            cands.append(url)
            continue

        # Excel MIME types
        if mime and (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" in mime
            or "application/vnd.ms-excel" in mime
        ):
            cands.append(url)

    # de-dupe while preserving order
    seen = set()
    out = []
    for u in cands:
        if u in seen:
            continue
        seen.add(u)
        out.append(u)
    return out


def _find_clickables_by_keywords(driver: webdriver.Chrome, keywords: list[str]):
    """Find clickable elements (a/button/div/span) containing any keyword in visible text."""
    # XPath: case-insensitive by translating to lowercase
    lower = "abcdefghijklmnopqrstuvwxyz"
    upper = lower.upper()

    parts = []
    for kw in keywords:
        kw_l = kw.lower()
        parts.append(
            f"contains(translate(normalize-space(string(.)), '{upper}', '{lower}'), '{kw_l}')"
        )

    cond = " or ".join(parts)

    xpath = (
        "//a[" + cond + "]"
        " | //button[" + cond + "]"
        " | //*[@role='button'][" + cond + "]"
        " | //span[" + cond + "]"
        " | //div[" + cond + "]"
    )

    try:
        els = driver.find_elements(By.XPATH, xpath)
    except Exception:
        return []

    # Filter to those that are displayed/enabled-ish
    out = []
    for el in els:
        try:
            if el.is_displayed():
                out.append(el)
        except Exception:
            continue

    return out


def _discover_xlsx_url(driver: webdriver.Chrome, unitid: str, debug: bool = False) -> str:
    """Best-effort: find an XLSX export link on the institution profile page.

    Strategy:
    1) Look for direct <a href="...xlsx"> links.
    2) If none, attempt to click Export/Excel/Download controls and inspect network activity for XLSX-ish URLs
       or Excel MIME types (JS-triggered downloads).
    """
    url = NCES_PROFILE_URL_ANCHOR.format(unitid)
    driver.get(url)
    time.sleep(MIN_SLEEP_AFTER_GET_SEC)

    wait = WebDriverWait(driver, WAIT_TIMEOUT_SEC)
    wait.until(lambda d: len(d.find_element(By.TAG_NAME, "body").text.strip()) > 200)

    if debug:
        print(f"\n[XLSX-DBG] unitid={unitid} loaded_url={driver.current_url}")

    anchors = driver.find_elements(By.TAG_NAME, "a")

    if debug:
        hrefs = [(a.get_attribute("href") or "").strip() for a in anchors]
        href_count = sum(1 for h in hrefs if h)
        downloadish = sum(1 for h in hrefs if h and "download" in h.lower())
        print(f"[XLSX-DBG] anchors={len(anchors)} with_href={href_count} href_contains_download={downloadish}")

        shown = 0
        for a in anchors:
            if shown >= 5:
                break
            txt = (a.text or "").strip()
            href = (a.get_attribute("href") or "").strip()
            if txt or href:
                print(f"[XLSX-DBG] sample_anchor text={txt[:120]!r} href={href[:200]!r}")
                shown += 1

    # 1) Prefer direct .xlsx hrefs
    for a in anchors:
        href = (a.get_attribute("href") or "").strip()
        if ".xlsx" in href.lower():
            if debug:
                print(f"[XLSX-DBG] FOUND direct .xlsx href={href}")
            return href

    # 2) JS-triggered export: click and capture network
    if XLSX_CLICK_AND_CAPTURE:
        if debug:
            print(f"[XLSX-DBG] no direct .xlsx links; attempting click+capture (window={XLSX_CAPTURE_WINDOW_SEC}s)")

        # Clear any prior perf logs
        _ = _drain_performance_logs(driver)

        keywords = ["export", "excel", "download", "xlsx"]
        clickables = _find_clickables_by_keywords(driver, keywords)

        if debug:
            print(f"[XLSX-DBG] clickables_found={len(clickables)}")
            for el in clickables[:8]:
                try:
                    txt = (el.text or "").strip().replace("\n", " ")
                    tag = el.tag_name
                    print(f"[XLSX-DBG] clickable tag={tag!r} text={txt[:140]!r}")
                except Exception:
                    continue

        # Try clicking a few candidates and see what network calls occur
        for idx, el in enumerate(clickables[:6]):
            try:
                if debug:
                    print(f"[XLSX-DBG] clicking candidate {idx+1}/{min(len(clickables),6)}")

                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
                time.sleep(0.15)

                # attempt click
                try:
                    el.click()
                except Exception:
                    driver.execute_script("arguments[0].click();", el)

                # wait a bit and capture perf logs
                t_end = time.time() + XLSX_CAPTURE_WINDOW_SEC
                all_logs = []
                while time.time() < t_end:
                    time.sleep(0.35)
                    all_logs.extend(_drain_performance_logs(driver))

                cands = _extract_xlsx_like_urls_from_perf(all_logs)
                if debug:
                    print(f"[XLSX-DBG] network_candidates_after_click={len(cands)}")
                    for u in cands[:10]:
                        print(f"  - {u}")

                # Pick the best candidate
                for u in cands:
                    if ".xlsx" in u.lower():
                        return u

                if cands:
                    # Return first plausible export URL even if it doesn't end in .xlsx
                    return cands[0]

            except Exception as e:
                if debug:
                    print(f"[XLSX-DBG] click attempt failed: {e}")
                continue

    # Extra debug: page_source keyword probes
    if debug:
        src = driver.page_source or ""
        src_l = src.lower()
        for tok in ("xlsx", "excel", "export", "download"):
            print(f"[XLSX-DBG] token_in_source {tok!r} -> {tok in src_l}")

    return ""


# ----------------------------
# Helpers: HTML -> tables
# ----------------------------

_WS = re.compile(r"\s+")


def _norm(s: str) -> str:
    return _WS.sub(" ", (s or "").strip())


def _find_table_near_heading(soup: BeautifulSoup, heading_substring: str) -> Optional[BeautifulSoup]:
    """Find the first <table> that appears after an element containing heading_substring."""
    heading_substring_l = heading_substring.lower()

    # Search for any element with matching visible text.
    el = soup.find(string=lambda t: t and heading_substring_l in t.lower())
    if not el:
        return None

    container = el.parent

    # Walk forward in the DOM looking for a table.
    nxt = container
    for _ in range(0, 40):
        if not nxt:
            break
        if getattr(nxt, "name", None) == "table":
            return nxt
        # try next elements
        nxt = nxt.find_next()
        if getattr(nxt, "name", None) == "table":
            return nxt

    return None


def _table_to_df(table_tag) -> Optional[pd.DataFrame]:
    """Convert a bs4 <table> tag to a DataFrame using pandas.read_html."""
    if table_tag is None:
        return None

    try:
        dfs = pd.read_html(str(table_tag))
    except Exception:
        return None

    if not dfs:
        return None
    df = dfs[0]
    # normalize column names
    df.columns = [str(c) for c in df.columns]
    return df


def _value_from_df(df: pd.DataFrame, row_key: str, col_key: str) -> str:
    """Best-effort: find df cell where a row contains row_key and column contains col_key."""
    if df is None or df.empty:
        return ""

    row_key_l = row_key.lower()
    col_key_l = col_key.lower()

    # identify column
    col_candidates = [c for c in df.columns if col_key_l in str(c).lower()]
    if not col_candidates:
        col_candidates = [df.columns[-1]]  # fallback to last column
    col = col_candidates[0]

    # identify row
    row_idx = None
    for i in range(len(df)):
        row_text = " ".join([str(x) for x in df.iloc[i].tolist()])
        if row_key_l in row_text.lower():
            row_idx = i
            break

    if row_idx is None:
        return ""

    val = df.iloc[row_idx][col]
    if pd.isna(val):
        return ""
    return _norm(str(val))


# ----------------------------
# Extractors
# ----------------------------


def extract_tuition_2024_25(soup: BeautifulSoup) -> Tuple[str, str]:
    """Return (undergrad, grad) tuition+fees for AY 2024-25 as strings."""
    # The heading on NCES pages usually contains this phrase.
    table = _find_table_near_heading(soup, "Tuition and required fees for full-time students")
    df = _table_to_df(table)
    if df is None:
        return "", ""

    # Try to locate the specific academic year column.
    # Columns vary; we attempt to choose the one that includes 2024-25.
    year_col = None
    for c in df.columns:
        if "2024" in str(c) and "25" in str(c):
            year_col = str(c)
            break
    if year_col is None:
        # sometimes year is in first row headers; just fall back to last column
        year_col = str(df.columns[-1])

    ug = _value_from_df(df, "Undergraduate", year_col)
    grad = _value_from_df(df, "Graduate", year_col)

    return ug, grad


def extract_enrollment_gender_totals(soup: BeautifulSoup) -> Tuple[str, str, str]:
    """Return (total, men, women) from the 'Enrollment by gender...' table."""
    table = _find_table_near_heading(soup, "Enrollment by gender")
    df = _table_to_df(table)
    if df is None:
        return "", "", ""

    # This table can be multi-level; common pattern has rows for Men/Women/Total.
    # We'll look for a 'Total' column if present; otherwise use last numeric column.
    total = _value_from_df(df, "Total", "Total")
    men = _value_from_df(df, "Men", "Total")
    women = _value_from_df(df, "Women", "Total")

    # If those are empty, try other likely keys.
    if not total:
        total = _value_from_df(df, "Total", df.columns[-1])
    if not men:
        men = _value_from_df(df, "Men", df.columns[-1])
    if not women:
        women = _value_from_df(df, "Women", df.columns[-1])

    return total, men, women


RACE_KEYS = [
    "American Indian or Alaska Native",
    "Asian",
    "Black or African American",
    "Hispanic",
    "Native Hawaiian or Other Pacific Islander",
    "White",
    "Two or more races",
    "Race/ethnicity unknown",
    "U.S. Nonresident",
]

# List of new value columns in stable order (tuition, enrollment, pct_<race>)
NEW_VALUE_COLS = (
    [
        "nces_profile_xlsx_downloaded",
        "nces_profile_xlsx_filename",
        "nces_profile_xlsx_dir",
        "nces_profile_xlsx_parse_ok",
        "nces_profile_xlsx_error",
        "tuition_fees_ug_2024_25",
        "tuition_fees_grad_2024_25",
        "enrollment_total",
        "enrollment_men",
        "enrollment_women",
    ]
    + [f"pct_{k}" for k in RACE_KEYS]
)

# ----------------------------
# XLSX download and extraction
# ----------------------------

SECTION_TUITION = "tuition and required fees for full-time students"
SECTION_ENROLLMENT = "Enrollment by gender, student level, and full- and part-time status"
SECTION_RACE = "Percent of all students enrolled, by race/ethnicity"


def _download_profile_xlsx(driver: webdriver.Chrome, unitid: str, debug: bool = False) -> Optional[Path]:
    """Click 'Download Profile' and wait for an .xlsx to appear in a unitid-specific download folder.

    This avoids filename collisions across long batch runs.
    """
    url = NCES_PROFILE_URL_ANCHOR.format(unitid)
    driver.get(url)
    time.sleep(MIN_SLEEP_AFTER_GET_SEC)

    wait = WebDriverWait(driver, WAIT_TIMEOUT_SEC)
    wait.until(lambda d: len(d.find_element(By.TAG_NAME, "body").text.strip()) > 200)

    # Create a unique download folder per unitid + timestamp (ms)
    base = _ensure_download_dir()
    unit_dir = base / unitid / str(int(time.time() * 1000))
    unit_dir.mkdir(parents=True, exist_ok=True)

    # Point Chrome downloads to THIS folder for THIS click
    try:
        driver.execute_cdp_cmd(
            "Page.setDownloadBehavior",
            {"behavior": "allow", "downloadPath": str(unit_dir)},
        )
    except Exception:
        try:
            driver.execute_cdp_cmd(
                "Browser.setDownloadBehavior",
                {"behavior": "allow", "downloadPath": str(unit_dir)},
            )
        except Exception:
            pass

    # Prefer an element that explicitly says "Download Profile"
    lower = "abcdefghijklmnopqrstuvwxyz"
    upper = lower.upper()
    cond = "contains(translate(normalize-space(string(.)), '%s', '%s'), 'download profile')" % (upper, lower)

    xpath = (
        "//a[" + cond + "]"
        " | //button[" + cond + "]"
        " | //*[@role='button'][" + cond + "]"
    )

    els = driver.find_elements(By.XPATH, xpath)
    if debug:
        print(f"[DL-DBG] unitid={unitid} download_profile_matches={len(els)} folder={unit_dir}")

    if not els:
        # fallback: any element containing 'download' and 'profile'
        cond2 = (
            "contains(translate(normalize-space(string(.)), '%s', '%s'), 'download') "
            "and contains(translate(normalize-space(string(.)), '%s', '%s'), 'profile')"
        ) % (upper, lower, upper, lower)
        xpath2 = (
            "//a[" + cond2 + "]"
            " | //button[" + cond2 + "]"
            " | //*[@role='button'][" + cond2 + "]"
        )
        els = driver.find_elements(By.XPATH, xpath2)
        if debug:
            print(f"[DL-DBG] unitid={unitid} download+profile_fallback_matches={len(els)}")

    if not els:
        return None

    # Click the first visible match
    clicked = False
    for el in els[:3]:
        try:
            if not el.is_displayed():
                continue
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
            time.sleep(0.15)
            try:
                el.click()
            except Exception:
                driver.execute_script("arguments[0].click();", el)
            clicked = True
            break
        except Exception:
            continue

    if not clicked:
        return None

    # Wait for any .xlsx in that folder (ignore partials)
    t_end = time.time() + DOWNLOAD_WAIT_SEC
    while time.time() < t_end:
        time.sleep(0.35)

        if list(unit_dir.glob("*.crdownload")) or list(unit_dir.glob("*.tmp")):
            continue

        xlsxs = sorted(unit_dir.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
        if xlsxs:
            return xlsxs[0]

    return None
# XLSX helpers

def _xlsx_find_cell(ws, needle: str) -> Optional[Tuple[int, int]]:
    """Find the first cell whose text contains needle (case-insensitive)."""
    needle_l = needle.lower()
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if v is None:
                continue
            if needle_l in str(v).lower():
                return (cell.row, cell.column)
    return None


def _xlsx_block_to_df(ws, start_row: int, start_col: int, max_rows: int = 80, max_cols: int = 40) -> pd.DataFrame:
    """Read a rectangular block of worksheet cells into a raw DataFrame."""
    data = []
    for r in range(start_row, start_row + max_rows):
        row = []
        for c in range(start_col, start_col + max_cols):
            row.append(ws.cell(row=r, column=c).value)
        data.append(row)
    return pd.DataFrame(data)



# --- Robust XLSX parsing helpers ---
def _cell_str(x) -> str:
    return "" if x is None else str(x).strip()


def _norm_lower(x) -> str:
    return _cell_str(x).lower()


def _is_numericish(v) -> bool:
    """Accept ints/floats or strings that look like counts/percents/currency (reject labels)."""
    if v is None:
        return False
    if isinstance(v, (int, float)):
        return True
    s = str(v).strip()
    if s == "":
        return False
    return re.fullmatch(r"[\$]?\s*\d[\d,]*(\.\d+)?\s*%?", s) is not None


def _find_header_map(block: pd.DataFrame, required_headers: list[str], search_rows: int = 15) -> Tuple[Optional[int], Dict[str, int]]:
    """Find a header row containing all required headers (case-insensitive substring match).

    Returns: (header_row_index, mapping of required_header_lower -> col_index)
    """
    req_l = [h.lower() for h in required_headers]

    for r in range(min(search_rows, len(block))):
        row = [_norm_lower(x) for x in block.iloc[r].tolist()]
        mapping: Dict[str, int] = {}

        for h in req_l:
            for c, cell in enumerate(row):
                if h in cell:
                    mapping[h] = c
                    break

        if all(h in mapping for h in req_l):
            return r, mapping

    return None, {}


def _first_nonempty_cell_value(row_vals: list) -> str:
    for x in row_vals:
        s = _cell_str(x)
        if s != "":
            return s
    return ""


def _find_row_by_first_cell(block: pd.DataFrame, start_row: int, label: str, max_scan: int = 120) -> Optional[int]:
    """Find the first row where the first non-empty cell equals label (case-insensitive)."""
    label_l = label.lower()
    end = min(len(block), start_row + max_scan)

    for r in range(start_row, end):
        first = _first_nonempty_cell_value(block.iloc[r].tolist())
        if first.lower() == label_l:
            return r

    return None


def _get_cell_norm(block: pd.DataFrame, r: int, c: int) -> str:
    if r is None:
        return ""
    if r < 0 or r >= len(block) or c < 0 or c >= len(block.columns):
        return ""
    v = block.iloc[r, c]
    if v is None:
        return ""
    return _norm(str(v))


def extract_all_fields_from_xlsx(xlsx_path: Path, unitid: Optional[str] = None) -> Dict[str, str]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    # Validate file matches the expected institution.
    if unitid:
        want = f"{unitid} -"
        found = False
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(values_only=True):
                for v in row:
                    if v is None:
                        continue
                    if want in str(v):
                        found = True
                        break
                if found:
                    break
            if found:
                break
        if not found:
            raise ValueError(f"Downloaded XLSX does not appear to contain unitid '{unitid}'")

    def find_section_block(section_name: str) -> Optional[pd.DataFrame]:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            loc = _xlsx_find_cell(ws, section_name)
            if not loc:
                continue
            r, c = loc
            # Block begins one row below section title
            return _xlsx_block_to_df(ws, r + 1, c, max_rows=140, max_cols=70)
        return None

    out: Dict[str, str] = {}

    # --------------------
    # Tuition section
    # --------------------
    tuition_block = find_section_block(SECTION_TUITION)
    ug = grad = ""
    if tuition_block is not None and not tuition_block.empty:
        hr, hm = _find_header_map(
            tuition_block,
            required_headers=["Level of student", "Tuition and required fees"],
            search_rows=20,
        )
        if hr is not None:
            tuition_col = hm["tuition and required fees"]

            # Search for rows below header
            sub = tuition_block.iloc[hr + 1 :].reset_index(drop=True)
            ug_r = _find_row_by_first_cell(sub, 0, "Undergraduate")
            gr_r = _find_row_by_first_cell(sub, 0, "Graduate")

            if ug_r is not None:
                v = sub.iloc[ug_r, tuition_col]
                ug = _norm(str(v)) if _is_numericish(v) else ""
            if gr_r is not None:
                v = sub.iloc[gr_r, tuition_col]
                grad = _norm(str(v)) if _is_numericish(v) else ""

    out["tuition_fees_ug_2024_25"] = ug
    out["tuition_fees_grad_2024_25"] = grad

    # --------------------
    # Enrollment section
    # --------------------
    enroll_block = find_section_block(SECTION_ENROLLMENT)
    total = men = women = ""
    if enroll_block is not None and not enroll_block.empty:
        hr, hm = _find_header_map(enroll_block, required_headers=["Total", "Men", "Women"], search_rows=25)
        if hr is not None:
            total_c = hm["total"]
            men_c = hm["men"]
            women_c = hm["women"]

            sub = enroll_block.iloc[hr + 1 :].reset_index(drop=True)
            all_r = _find_row_by_first_cell(sub, 0, "All students")
            if all_r is not None:
                v_total = sub.iloc[all_r, total_c]
                v_men = sub.iloc[all_r, men_c]
                v_women = sub.iloc[all_r, women_c]

                total = _norm(str(v_total)) if _is_numericish(v_total) else ""
                men = _norm(str(v_men)) if _is_numericish(v_men) else ""
                women = _norm(str(v_women)) if _is_numericish(v_women) else ""

    out["enrollment_total"] = total
    out["enrollment_men"] = men
    out["enrollment_women"] = women

    # --------------------
    # Race/ethnicity percentages
    # --------------------
    race_block = find_section_block(SECTION_RACE)
    if race_block is not None and not race_block.empty:
        hr, hm = _find_header_map(race_block, required_headers=RACE_KEYS, search_rows=25)
        if hr is not None:
            sub = race_block.iloc[hr + 1 :].reset_index(drop=True)
            data_r = _find_row_by_first_cell(sub, 0, "Enrollment by race/ethnicity")
            if data_r is not None:
                for k in RACE_KEYS:
                    c = hm[k.lower()]
                    v = sub.iloc[data_r, c]
                    out[f"pct_{k}"] = _norm(str(v)) if _is_numericish(v) else ""
            else:
                for k in RACE_KEYS:
                    out[f"pct_{k}"] = ""
        else:
            for k in RACE_KEYS:
                out[f"pct_{k}"] = ""
    else:
        for k in RACE_KEYS:
            out[f"pct_{k}"] = ""

    return out


def extract_race_ethnicity_percentages(soup: BeautifulSoup) -> Dict[str, str]:
    """Return mapping of race/ethnicity label -> percentage string."""
    table = _find_table_near_heading(soup, "Percent of all students enrolled")
    df = _table_to_df(table)
    out: Dict[str, str] = {k: "" for k in RACE_KEYS}

    if df is None:
        return out

    # Determine which column likely holds percentages.
    pct_col = None
    for c in df.columns:
        cl = str(c).lower()
        if "percent" in cl or "%" in cl:
            pct_col = str(c)
            break
    if pct_col is None:
        pct_col = str(df.columns[-1])

    for k in RACE_KEYS:
        out[k] = _value_from_df(df, k, pct_col)

    return out


def extract_all_fields_from_html(html: str) -> Dict[str, str]:
    soup = BeautifulSoup(html, "html.parser")

    ug, grad = extract_tuition_2024_25(soup)
    total, men, women = extract_enrollment_gender_totals(soup)
    race = extract_race_ethnicity_percentages(soup)

    out: Dict[str, str] = {
        "tuition_fees_ug_2024_25": ug,
        "tuition_fees_grad_2024_25": grad,
        "enrollment_total": total,
        "enrollment_men": men,
        "enrollment_women": women,
    }

    for k in RACE_KEYS:
        out[f"pct_{k}"] = race.get(k, "")

    return out


# ----------------------------
# Runner
# ----------------------------


def run(test_n: int = TEST_N) -> pd.DataFrame:
    df_full = pd.read_csv(INPUT_PATH, dtype=str, keep_default_na=False)

    if UNITID_COL not in df_full.columns:
        raise ValueError(f"Missing required column '{UNITID_COL}' in {INPUT_PATH}")
    if NAME_COL not in df_full.columns:
        raise ValueError(f"Missing required column '{NAME_COL}' in {INPUT_PATH}")

    # For fetching, sample the first N UNIQUE unitids (testing knob).
    # If test_n == 0, run over ALL unique unitids.
    uniq = df_full[[UNITID_COL, NAME_COL]].drop_duplicates(subset=[UNITID_COL]).copy()
    if int(test_n) == 0:
        test_df = uniq
    else:
        test_df = uniq.head(int(test_n)).copy()

    rows = []
    driver = make_driver()
    try:
        for _, r in test_df.iterrows():
            unitid = str(r[UNITID_COL]).strip()
            name = str(r[NAME_COL]).strip()

            rec: Dict[str, str] = {"unitid": unitid, "name": name}

            # Default empty new fields so merge always has columns
            for c in NEW_VALUE_COLS:
                rec.setdefault(c, "")

            if not unitid:
                rows.append(rec)
                continue

            try:
                rec["nces_profile_xlsx_downloaded"] = "0"
                rec["nces_profile_xlsx_filename"] = ""
                rec["nces_profile_xlsx_dir"] = ""
                rec["nces_profile_xlsx_parse_ok"] = "0"
                rec["nces_profile_xlsx_error"] = ""

                used_xlsx = False
                parsed_ok = False

                if USE_XLSX_EXTRACTION:
                    xlsx_path = _download_profile_xlsx(driver, unitid, debug=DEBUG_XLSX_DISCOVERY)
                    if xlsx_path is not None and xlsx_path.exists():
                        rec["nces_profile_xlsx_downloaded"] = "1"
                        rec["nces_profile_xlsx_filename"] = xlsx_path.name
                        rec["nces_profile_xlsx_dir"] = str(xlsx_path.parent)
                        try:
                            rec.update(extract_all_fields_from_xlsx(xlsx_path, unitid=unitid))
                            parsed_ok = True
                            rec["nces_profile_xlsx_parse_ok"] = "1"
                        except Exception as xe:
                            rec["nces_profile_xlsx_error"] = str(xe)[:300]
                        used_xlsx = True

                        # Only delete XLSX if requested AND parsing succeeded
                        if DELETE_DOWNLOADED_XLSX and parsed_ok:
                            try:
                                xlsx_path.unlink()
                            except Exception:
                                pass

                # Fallback to HTML if we didn't use XLSX or if extraction looks empty
                def _all_empty(keys: list[str]) -> bool:
                    return all((rec.get(k, "") or "").strip() == "" for k in keys)

                critical_keys = [
                    "tuition_fees_ug_2024_25",
                    "tuition_fees_grad_2024_25",
                    "enrollment_total",
                    "enrollment_men",
                    "enrollment_women",
                ]

                if (not used_xlsx) or _all_empty(critical_keys):
                    html = _load_html(driver, unitid)
                    rec.update(extract_all_fields_from_html(html))

            except Exception as e:
                print(f"[WARN] unitid={unitid}: failed to fetch/parse ({e})")

            rows.append(rec)
            time.sleep(SLEEP_BETWEEN_REQUESTS_SEC)

    finally:
        driver.quit()

    unit_rows_df = pd.DataFrame(rows)

    # Ensure one row per unitid for merge
    if not unit_rows_df.empty:
        unit_rows_df = unit_rows_df.drop_duplicates(subset=[UNITID_COL]).copy()

    # Avoid clobbering the input file's name column
    if NAME_COL in unit_rows_df.columns:
        unit_rows_df = unit_rows_df.drop(columns=[NAME_COL], errors="ignore")

    merged = df_full.merge(unit_rows_df, on=UNITID_COL, how="left")

    # Ensure all new columns exist (blank if unitid not fetched)
    for c in NEW_VALUE_COLS:
        if c not in merged.columns:
            merged[c] = ""

    merged = merged.fillna("")

    output_path = INPUT_PATH.with_name(INPUT_PATH.stem + OUTPUT_SUFFIX + ".csv")
    merged.to_csv(output_path, index=False)

    # Short confirmation + small preview
    pd.set_option("display.max_columns", 200)
    pd.set_option("display.width", 220)

    preview_cols = [UNITID_COL] + [c for c in NEW_VALUE_COLS if c in merged.columns]
    print(f"Wrote: {output_path}")
    print(merged.loc[:, preview_cols].head(5))

    return merged


def _load_html(driver: webdriver.Chrome, unitid: str, post_get_sleep_sec: float = MIN_SLEEP_AFTER_GET_SEC) -> str:
    url = NCES_PROFILE_URL.format(unitid)
    driver.get(url)
    time.sleep(post_get_sleep_sec)

    wait = WebDriverWait(driver, WAIT_TIMEOUT_SEC)

    # IPEDS profile pages are JS-rendered; wait until the BODY has some real content.
    wait.until(lambda d: len(d.find_element(By.TAG_NAME, "body").text.strip()) > 200)

    return driver.page_source


if __name__ == "__main__":
    run(TEST_N)